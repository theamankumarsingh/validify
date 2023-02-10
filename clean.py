import pandas as pd
import openpyxl
import re
import requests

header = {
    'user-agent': 'Mozilla/5.0 (X11; Ubuntu; Linux x86_64; rv:15.0) Gecko/20100101 Firefox/15.0.1',
}

def dict_to_workbook(working_data_dict,data_problem_dict,workbook_name):
    wb = openpyxl.Workbook()
    sheet = wb['Sheet']
    keyss=list(working_data_dict.keys())
    keyss.sort()
    row=2
    #Headings
    sheet['A1']="Key No."
    sheet['B1']="Website"
    sheet['C1']="Problem"
    for key in keyss:
        sheet['A'+str(row)]=key
        sheet['B'+str(row)]=working_data_dict[key]
        if data_problem_dict[key]==():
            sheet['C'+str(row)]="autofix"
        else:
            sheet['C'+str(row)]=str(data_problem_dict[key])
        row+=1
    wb.save(workbook_name)

def workbook_to_dict(workbook_name):
    data_dict={}
    wb=openpyxl.load_workbook(workbook_name)
    sheet=wb['Sheet']
    for row in range(2,sheet.max_row+1):
        data_dict[sheet['A'+str(row)].value]=sheet['B'+str(row)].value
    return data_dict

def start_manual(working_data_dict,data_problem_dict,workbook_name,skip=True):
    if skip:
        return working_data_dict
    print("Starting manual cleaning ...")
    dict_to_workbook(working_data_dict,data_problem_dict,workbook_name)
    input("Please edit the file "+workbook_name+" and save it before proceeding. Press Enter to continue ...")
    return workbook_to_dict(workbook_name)

#auto-clean functions
def clean_url(url):
    pattern = r"[^\w]+"
    match = re.search(pattern, url)
    if not match:
        return "","",domain
    prefix=url[:match.start()]
    domain=url[match.end():]
    symbol=match.group()
    return prefix,symbol,domain

def start_autofix(working_data_dict,data_problem_dict,skip=True):
    if skip:
        return working_data_dict,data_problem_dict
    print("Starting autofix ...")
    for key in working_data_dict.keys():
        prefix,symbol,domain=clean_url(working_data_dict[key])
        if symbol!="" or domain!="":
            best_guess=fix_url(prefix,symbol,domain)
            if best_guess!=prefix+symbol+domain:
                working_data_dict[key]=best_guess
                data_problem_dict[key]=()
    return working_data_dict,data_problem_dict

def fix_url(prefix,symbol,domain):
    url_1="https://"+domain
    url_2="http://"+domain
    try:
        response_1=requests.get(url_1,headers=header)
        stat_code_1=response_1.status_code
    except Exception as exception:
        stat_code_1=0
    try:
        response_2=requests.get(url_2,headers=header)
        stat_code_2=response_2.status_code
    except Exception as exception:
        stat_code_2=0
    
    if stat_code_1==200:
        return url_1
    elif stat_code_2==200:
        return url_2
    else:
        return prefix+symbol+domain