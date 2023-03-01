import multiprocessing
import sys
import openpyxl
import preparation
import subprocess
import numpy as np
import os

def launch_script_with_terminal(wb_name,code=0):
    subprocess.Popen(["gnome-terminal","--","python3","script.py",wb_name,code])

if len(sys.argv) < 3:
    print("Usage: python3 script.py <workbook_name> <processes> <operation code(optional)>")
    sys.exit(1)
else:
    workbook_name=sys.argv[1]
process=int(sys.argv[2])
if process < 1:
    print("Error: process count must be greater than 0")
    sys.exit(1)
code=0  #code 0 for autofix only, code 1 for autofix and manual, code 2 for code 0 with check and code 3 for code 1 with check
if len(sys.argv) >= 4:
    code=int(sys.argv[3])
    if code not in [0,1,2,3]:
        code=0
workbook_ext=".xlsx"
if workbook_name[-5:] == ".xlsx":
    workbook_name=workbook_name[:-5]

logo='''
                                                  
                                                  
                 ######(   (#####    .%%          
             (((                   ##    %%       
          ((         ((         ,##   %%          
        #(                    ##          %       
      ##       .((   (/     ##   ##        #(     
     ##      #( ((    (     (    ##         ##    
    /#    .##    ((   ((   ((   ##           #    
        ##        #    (( ((    (       ##   ##   
     .%%   ##      #    ( (    (      ##     /#   
   %%              ##   /(    ((   ((        ##   
 &&                 ##       ((              #    
    &%               #*     /(      ((      ##    
       #        %%*   #     #              (,     
        &&    ,%         ##              ((       
          &&           %%              ((         
             &&*    %%    %        ###            
                  %%   %%    ####                 
               /&    %%                           
'''

print(logo)

#read sheet from workbook
print("Preparing file "+workbook_name+workbook_ext+" for execution ...")
try:
    wb=openpyxl.load_workbook(workbook_name+workbook_ext)
except Exception as exception:
    print("Error: "+type(exception).__name__+" while reading workbook from file "+workbook_name+workbook_ext)
    sys.exit(1)
sheet=wb['Sheet']
dataset,data_dict=preparation.prepare(sheet,duplicates=False)
wb.close()
dataset_t = np.array_split(dataset, process)

processes=[]
for i in range(process):
    dataset_n=list(dataset_t[i])
    wb_new = openpyxl.Workbook()
    sheet_new = wb_new['Sheet']
    sheet_new['A1']="Name"
    sheet_new['B1']="Register Number"
    sheet_new['C1']="Mobile Number"
    sheet_new['D1']="Email-Id"
    sheet_new['E1']="WebURL"
    for row in range(1,len(dataset_n)+1):
        sheet_new['A'+str(row+1)]=dataset_n[row-1][1]
        sheet_new['B'+str(row+1)]=dataset_n[row-1][2]
        sheet_new['C'+str(row+1)]=dataset_n[row-1][3]
        sheet_new['D'+str(row+1)]=dataset_n[row-1][4]
        sheet_new['E'+str(row+1)]=dataset_n[row-1][5]
    wb_new.save(workbook_name+str(i)+workbook_ext)
    wb_new.close()
    processes.append(multiprocessing.Process(target=launch_script_with_terminal,args=(workbook_name+str(i)," "+str(code))))

#start all processes
print("Starting all processes ("+str(process)+") ...")
for p in processes:
    p.start()
input("Please wait for completion, then press enter to exit ...")

#join all worksheets
final_dataset=[]
print("Joining all worksheets ...")
for i in range(process):
    try:
        wb=openpyxl.load_workbook(workbook_name+str(i)+"_out"+workbook_ext)
    except Exception as exception:
        print("Error: "+type(exception).__name__+" while reading workbook from file "+workbook_name+str(i)+"_out"+workbook_ext)
        sys.exit(1)
    sheet=wb['Sheet']
    for row in range(2,sheet.max_row+1):
        final_dataset.append((sheet['A'+str(row)].value,sheet['B'+str(row)].value,sheet['C'+str(row)].value,sheet['D'+str(row)].value,sheet['E'+str(row)].value))
    wb.close()

#write final dataset to file
print("Writing to file "+workbook_name+"_output"+workbook_ext+" ...")
wb_new = openpyxl.Workbook()
sheet_new = wb_new['Sheet']
sheet_new['A1']="Name"
sheet_new['B1']="Register Number"
sheet_new['C1']="Mobile Number"
sheet_new['D1']="Email-Id"
sheet_new['E1']="WebURL"
for row in range(1,len(final_dataset)+1):
    sheet_new['A'+str(row+1)]=final_dataset[row-1][0]
    sheet_new['B'+str(row+1)]=final_dataset[row-1][1]
    sheet_new['C'+str(row+1)]=final_dataset[row-1][2]
    sheet_new['D'+str(row+1)]=final_dataset[row-1][3]
    sheet_new['E'+str(row+1)]=final_dataset[row-1][4]
wb_new.save(workbook_name+"_output"+workbook_ext)
wb_new.close()

#cleaning up
pwd=os.getcwd()
folder_path = pwd
files_to_keep = [workbook_name+workbook_ext,workbook_name+"_output"+workbook_ext]
for filename in os.listdir(folder_path):
    if filename.endswith(".xlsx") and filename not in files_to_keep:
        os.remove(os.path.join(folder_path, filename))

print("All done!")
sys.exit(0)