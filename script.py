#imports
import openpyxl
import preparation
import clean
import finalization
import sys

#read source workbook from command line
if len(sys.argv) < 2:
    print("Usage: python3 script.py <workbook_name>")
    sys.exit(1)
else:
    workbook_name=sys.argv[1]
workbook_ext=".xlsx"
if workbook_name[-5:] == ".xlsx":
    workbook_name=workbook_name[:-5]

#read sheet from workbook
print("Reading workbook from file "+workbook_name+workbook_ext+" ...")
try:
    wb=openpyxl.load_workbook(workbook_name+workbook_ext)
except Exception as exception:
    print("Error: "+type(exception).__name__+" while reading workbook from file "+workbook_name+workbook_ext)
    sys.exit(1)
sheet=wb['Sheet']

#pre-processing
dataset,data_dict=preparation.prepare(sheet,duplicates=True)
working_data_dict,data_problem_dict=preparation.filter_bad_data(data_dict)
backup_data_dict=working_data_dict.copy()

#cleaning
#autoclean
working_data_dict,data_problem_dict=clean.start_autofix(working_data_dict,data_problem_dict,skip=False)

#clean (manual)
working_data_dict=clean.start_manual(working_data_dict,data_problem_dict,workbook_name+"_edit"+workbook_ext,skip=True)

#post-processing
diff_res,problem_keys=finalization.diff(backup_data_dict,working_data_dict,data_problem_dict)
data_dict=finalization.change(working_data_dict,data_dict)
dataset=finalization.modify(dataset,data_dict)
dataset=finalization.remove_problem_keys(dataset,problem_keys)
test_res=finalization.check(dataset,skip=False)

#write sheet to workbook
print("Writing workbook to file "+workbook_name+"_out"+workbook_ext+" ...")
wb_new = openpyxl.Workbook()
sheet_new = wb_new['Sheet']
sheet_new['A1']="Name"
sheet_new['B1']="Register Number"
sheet_new['C1']="Mobile Number"
sheet_new['D1']="Email-Id"
sheet_new['E1']="WebURL"
for row in range(1,len(dataset)+1):
    sheet_new['A'+str(row+1)]=dataset[row-1][1]
    sheet_new['B'+str(row+1)]=dataset[row-1][2]
    sheet_new['C'+str(row+1)]=dataset[row-1][3]
    sheet_new['D'+str(row+1)]=dataset[row-1][4]
    sheet_new['E'+str(row+1)]=dataset[row-1][5]
if len(test_res):
    wb_new.create_sheet('Test')
    sheet_new = wb_new['Test']
    sheet_new['A1']="Website"
    sheet_new['B1']="Check"
    sheet_new['C1']="Code"
    for row in range(1,len(test_res)+1):
        sheet_new['A'+str(row+1)]=test_res[row-1][0]
        sheet_new['B'+str(row+1)]=test_res[row-1][1]
        sheet_new['C'+str(row+1)]=test_res[row-1][2]
if len(diff_res):
    wb_new.create_sheet('Diff')
    sheet_new = wb_new['Diff']
    sheet_new['A1']="Original Website"
    sheet_new['B1']="Changed Website"
    sheet_new['C1']="Fix"
    for row in range(1,len(diff_res)+1):
        sheet_new['A'+str(row+1)]=diff_res[row-1][0]
        sheet_new['B'+str(row+1)]=diff_res[row-1][1]
        sheet_new['C'+str(row+1)]=diff_res[row-1][2]
wb_new.save(workbook_name+"_out"+workbook_ext)
print("Done!")